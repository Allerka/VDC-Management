from decimal import *
from django.http import HttpResponseBadRequest
from django import forms
from django.template import RequestContext
from django.contrib.auth.models import User
from django.conf import settings
import django_excel as excel
from openpyxl.worksheet.datavalidation import DataValidation
import logging
import io
import shutil
from openpyxl import load_workbook
from .models import *

logger = logging.getLogger(__name__)

class UploadFileForm(forms.Form):
	# def __init__(self, request):
	file = forms.FileField()

	def upload(request):
		try:
			form = UploadFileForm(request.POST, request.FILES)
			if form.is_valid():
				filehandle = request.FILES['file']
				data = excel.ExcelMixin.get_sheet(filehandle)
				# logger.info(list(data.rows()))
				user = User.objects.get(username__icontains=request.headers['user'])
				corrected_data = convert(data)
				
				data_dict = {
					'r_name': str(user.first_name + " " + user.last_name),
					'r_email': str(user.email), 
					'r_phone': '', 
					'r_mobile': '', 
					't_name': '', 
					't_email': '',
					't_phone': '', 
					'year': corrected_data['year'], 
					'make': data['B5'], 
					'model': data['B6'], 
					'identifier': data['B7'], 
					'color': data['B8'], 
					'customer': data['B9'], 
					'test_name': data['B10'], 
					'vin': data['B11'], 
					'veh_type': corrected_data['veh_type'], 
					'f_pressure': round(Decimal(corrected_data['f_pressure']), 2), 
					'r_pressure': round(Decimal(corrected_data['r_pressure']), 2), 
					'fuel_type': data['B15'], 
					'fuel_type_other': data['B16'], 
					'fuel_cap': round(Decimal(corrected_data['fuel_cap']), 1), 
					'regulation': corrected_data['regulation'], 
					'cert_level': corrected_data['cert_level'], 
					'eng_family': data['B22'], 
					'emi_family': data['B23'], 
					'eng_out': corrected_data['eng_out'], 
					'cat_mid': corrected_data['cat_mid'], 
					'tail_post': corrected_data['tail_post'], 
					'marmon': corrected_data['marmon'], 
					'marmon_distance': round(Decimal(corrected_data['marmon_distance']), 2), 
					'usable_bat': round(Decimal(corrected_data['usable_bat']), 2), 
					'nom_bat': round(Decimal(corrected_data['nom_bat']), 2), 
					'all_range': round(Decimal(corrected_data['all_range']), 2), 
					'cscm': data['B32'], 
					'country': corrected_data['country'], 
					'mass_usa': corrected_data['mass_usa'], 
					'inertia': corrected_data['inertia'], 
					'test_mass_usa': round(Decimal(corrected_data['test_mass_usa']), 8), 
					'mass_ro': round(Decimal(corrected_data['mass_ro']), 2), 
					'mass_coc': round(Decimal(corrected_data['mass_coc']), 2), 
					'test_mass_eu': round(Decimal(corrected_data['test_mass_eu']), 6), 
					'coeff1': round(Decimal(corrected_data['coeff1']), 4), 
					'coeff2': round(Decimal(corrected_data['coeff2']), 6), 
					'coeff3': round(Decimal(corrected_data['coeff3']), 7), 
					'cold_co1': round(Decimal(corrected_data['cold_co1']), 2), 
					'cold_co2': round(Decimal(corrected_data['cold_co2']), 4), 
					'cold_co3': round(Decimal(corrected_data['cold_co3']), 5), 
					'wheelbase': data['B36'], 
					'dyno_mode': corrected_data['dyno_mode'], 
					'coastdown': corrected_data['coastdown'], 
					'dyno_roll': corrected_data['dyno_roll'], 
					'd_rings': corrected_data['d_rings'], 
					'front_hooks': corrected_data['front_hooks'], 
					'front_alt': data['B53'], 
					'rear_hooks': corrected_data['rear_hooks'], 
					'rear_alt': data['B55'], 
					'desc': data['B56'], 
					'charge_num': data['B58'], 
					'license': data['B59'], 
					'displacement': round(Decimal(corrected_data['displacement']), 1), 
					'cylinders': corrected_data['cylinders'], 
					'transmission': data['B62'], 
					'gears': corrected_data['gears'],
					'spreadsheet': filehandle,
					}
				# logger.info(data_dict)
				return data_dict
			else:
				return HttpResponseBadRequest()
		except Exception as e:
			logger.error(e)
			return e

	def download(data):
		filehandle = shutil.copy(settings.BASE_DIR + '/uploads/FEV_VDC_TestVehicleName_YYYYMMDD_V1.08.xlsx', settings.BASE_DIR + '/uploads/FEV_VDC_' + data['test_name'] + '_DRAFT.xlsx')
		workbook = load_workbook(filehandle)
		sheet = workbook.active
		field_list = ['', '', '', 'year', 'make', 'model', 'identifier', 'color', 'customer', 'test_name', 'vin', 'veh_type', 'f_pressure', 'r_pressure', 'fuel_type', 'fuel_type_other', 'fuel_cap', '', '', 'regulation', 'cert_level', 'eng_family', 'emi_family', 'eng_out', 'cat_mid', 'tail_post', 'marmon', 'marmon_distance', 'usable_bat', 'nom_bat', 'all_range', 'cscm', '', 'country', 'dyno_roll', 'wheelbase', 'mass_usa', 'inertia', 'test_mass_usa', 'mass_ro', 'mass_coc', 'test_mass_eu', 'coeff1', 'coeff2', 'coeff3', 'cold_co1', 'cold_co2', 'cold_co3', 'dyno_mode', 'coastdown', 'd_rings', 'front_hooks', 'front_alt', 'rear_hooks', 'rear_alt', 'desc', '', '', '', '', 'charge_num', 'license', 'displacement', 'cylinders', 'transmission', 'gears']
		valid_list = ['veh_type', 'fuel_type', 'regulation', 'cert_level', 'eng_out', 'cat_mid', 'tail_post', 'marmon', 'country', 'dyno_roll', 'inertia', 'dyno_mode', 'coastdown', 'd_rings', 'front_hooks', 'rear_hooks']
		for item, field in zip(range(1, 66), field_list):
			try:
				cell = sheet[item][1]
				if cell.value:
					pass
				else:
					if type(data[field]) == bool:
						sheet.cell(column=2, row=item, value=reverse_bool(field, data[field]))
					elif data[field] in ['NONE', 'None', 0, '0']: # keep cells blank instead of putting in default values from the form
						pass
					else:
						sheet.cell(column=2, row=item, value=data[field])
				if field in valid_list: # adding dropdown selectors
					list = get_formula(field)
					validator = DataValidation(type="list", formula1=list[0], allow_blank=True, operator='between')
					validator.promptTitle = list[1]
					validator.prompt = list[2] if list[2] else "Please select a " + list[1] + " from the dropdown list."
					if list[3]:
						validator.errorTitle = list[3]
						validator.error = list[4]
					else:
						validator.errorTitle = "Error"
						validator.error = "Please choose an option from the dropdown list."
					sheet.add_data_validation(validator)
					validator.add(cell)
			except KeyError: #field is blank and thus fails validation
				pass
			except Exception as e: 
				logger.exception(e)
		workbook.save(filehandle)
		return filehandle


# Parses certain Excel fields that don't conform to the standard registration's format
def convert(data):
	num_fields = {'year': 2020, 'f_pressure': 0, 'r_pressure': 0, 'fuel_cap': 0, 'marmon': 2, 'marmon_distance': 0, 'nom_bat': 0, 'usable_bat': 0, 'all_range': 0, 'cscm': 0, 'mass_usa': 0, 'test_mass_usa': 0, 'test_mass_eu': 0, 'mass_ro': 0, 'mass_coc': 0, 'coeff1': 0, 'coeff2': 0, 'coeff3': 0, 'cold_co1': 0, 'cold_co2': 0, 'cold_co3': 0, 'displacement': 0, 'cylinders': 0, 'gears': 0}

	for field, sheet in zip(num_fields, [data['B4'], data['B13'], data['B14'], data['B17'], data['B27'], data['B28'], data['B29'], data['B30'], data['B31'], data['B32'], data['B37'], data['B39'], data['B42'], data['B40'], data['B41'], data['B43'], data['B44'], data['B45'], data['B46'], data['B47'], data['B48'], data['B60'], data['B61'], data['B63']]):
		if sheet == None or sheet == '':
			num_fields[field] = num_fields[field]
		else:
			num_fields[field] = sheet

	# logger.info(num_fields)
	regulation = list_check(data['B20'], FieldChoices.get_options('regulation'))
	cert_level = list_check(data['B21'], FieldChoices.get_options('cert'))
	inertia = bool_check(data['B38'])
	eng_out = bool_check(data['B24'])
	cat_mid = bool_check(data['B25'])
	tail_post = bool_check(data['B26'])
	dyno_mode = list_check(data['B49'], FieldChoices.get_options('mode'))
	dyno_roll = list_check(data['B49'], FieldChoices.get_options('dyno_roll'))
	coastdown = list_check(data['B50'], FieldChoices.get_options('mode'))
	d_rings = bool_check(data['B51'])
	front_hooks = list_check(data['B52'], FieldChoices.get_options('hooks'))
	rear_hooks = list_check(data['B54'], FieldChoices.get_options('hooks'))
	country = list_check(data['B34'], FieldChoices.get_options('country'))
	veh_type = list_check(data['B12'], FieldChoices.get_options('veh_type'))

	return {'regulation': regulation, 'cert_level': cert_level, 'inertia': inertia, 'eng_out': eng_out, 'cat_mid': cat_mid, 'tail_post': tail_post, 'dyno_mode': dyno_mode, 'dyno_roll': dyno_roll, 'coastdown': coastdown, 'd_rings': d_rings, 'front_hooks': front_hooks, 'rear_hooks': rear_hooks, 'marmon_distance': num_fields['marmon_distance'], 'nom_bat': num_fields['nom_bat'], 'usable_bat': num_fields['usable_bat'], 'all_range': num_fields['all_range'], 'gears': num_fields['gears'], 'year': num_fields['year'], 'f_pressure': num_fields['f_pressure'], 'r_pressure': num_fields['r_pressure'], 'fuel_cap': num_fields['fuel_cap'], 'marmon': num_fields['marmon'], 'mass_usa': num_fields['mass_usa'], 'test_mass_usa': num_fields['test_mass_usa'], 'test_mass_eu': num_fields['test_mass_eu'], 'mass_ro': num_fields['mass_ro'], 'mass_coc': num_fields['mass_coc'], 'coeff1': num_fields['coeff1'], 'coeff2': num_fields['coeff2'], 'coeff3': num_fields['coeff3'], 'cold_co1': num_fields['cold_co1'], 'cold_co2': num_fields['cold_co2'], 'cold_co3': num_fields['cold_co3'], 'cylinders': num_fields['cylinders'], 'displacement': num_fields['displacement'], 'country': country, 'veh_type': veh_type}

# validates drop-down choices
def list_check(entry, list):
	for item in list:
		if item[1] == entry:
			return item[0]
	return 'NONE'

# validates yes/no choices
def bool_check(entry):
	if entry == 'Yes' or entry == 'Yes, labeled':
		return True
	elif entry == 'No' or entry == 'None' or 'No, see below': 
		return False
	else:
		raise ValueError

# converts bools to yes/no strings
def reverse_bool(field, data):
	if data == True:
		if field in ['inertia', 'd-rings']:
			return 'Yes'
		else:
			return 'Yes, labeled'
	elif data == False:
		if field == 'd_rings':
			return 'No, see below'
		else:
			return 'No'
	else:
		raise ValueError

# sets up the dropdown selections on a saved draft, along with their info popups
def get_formula(field):
	list = ['', '', '', '', '']
	if field == 'veh_type':
		list[0] = "=Resources!$O$2:$O$6"
		list[1] = "Vehicle Type"
	elif field == 'fuel_type':
		list[0] = "=Resources!$A$2:$A$17"
		list[1] = "Choose a fuel type from the list"
		list[2] = "If your fuel type is not listed, choose 'Other' and list your fuel type, then contact Alan Bedewi."
		list[3] = "Test Fuel Not Available"
		list[4] = "Choose 'Other' and list your fuel type, then contact Alan Bedewi."
	elif field == 'regulation':
		list[0] = "=Resources!$M$2:$M$10"
		list[1] = "Regulation"
	elif field == 'cert_level':
		list[0] = "=Resources!$N$2:$N$53"
		list[1] = "Certification Level"
	elif field == 'eng_out':
		list[0] = "=Resources!$E$2:$E$3"
		list[1] = "Engine Out/Feedgas/Precat Probe"
		list[2] = "If the vehicle has an Engine Out/Feedgas/Precat Probe installed for modal measurements, selecct Yes. If not, select No."
	elif field == 'cat_mid':
		list[0] = "=Resources!$E$2:$E$3"
		list[1] = "Catalyst Midbed Probe"
		list[2] = "If the vehcile has a Catalyst Midbed Probe installed for modal measurements, select Yes. If not, select No."
	elif field == 'tail_post':
		list[0] = "=Resources!$E$2:$E$3"
		list[1] = "Tailpipe/Postcat Probe"
		list[2] = "If the vehicle has a Tailpipe/Postcat Probe installed for modal measurements, select Yes. If not, select NO."
	elif field == 'marmon':
		list[0] = "=Resources!$H$2:$H$5"
		list[1] = "Marmon Flange Size"
		list[2] = "Select the marmon flange size from the drop down. Make sure your exhaust tube has a 1/8 to 1/4 inch overhang past the marmon flange to allow for exhaust gaskets."
	elif field == 'country':
		list[0] = "=Resources!$P$2:$P$7"
		list[1] = "Country"
		list[2] = "Choose a country designation from the drop-down menu to get the appropriate units."
	elif field == 'dyno_roll':
		list[0] = "=Resources!$B$2:$B$4"
		list[1] = "Dyno Roll Configuration"
		list[2] = "Choose a dyno roll configuration: FWD = Front wheels spinning only, RWD = Rear wheels spinning only, 4WD = all 4 wheels spinning (this can be an AWD vehicle, or a 2WD vehicle with no dyno mode)."
	elif field == 'inertia':
		list[0] = "=Resources!$C$2:$C$3"
		list[1] = "Inertia Class"
		list[2] = "Should the inertia class table be used instead of the vehicle's exact mass?"
	elif field == 'd_rings':
		list[0] = "=Resources!$G$2:$G$3"
		list[1] = "D-Rings (4 Locations)"
		list[2] = "If the vehicle is equipped with four D-Rings, choose Yes. If not, choose No and move on to Tow Hook options."
	elif field == 'dyno_mode':
		list[0] = "=Resources!$F$2:$F$4"
		list[1] = "Dyno Mode"
		list[2] = "If the vehicle has a dyno mode, select Yes and attach the procedure to this form. If not, select No."
	elif field == 'coastdown':
		list[0] = "=Resources!$F$2:$F$4"
		list[1] = "Neutral Coastdown Mode"
		list[2] = "If the vehicle has a neutral coastdown mode, select Yes and attach the procedure to this form. If not, select No."
	elif field == 'front_hooks':
		list[0] = "=Resources!$D$2:$D$5"
		list[1] = "Tow Hook(s) - Front"
		list[2] = "If the vehcile has a tow hook installed in the front, select Yes. If not, select No and specify the front tie down location for dyno installation."
	elif field == 'rear_hooks':
		list[0] = "=Resources!$D$2:$D$5"
		list[1] = "Tow Hook(s) - Rear"
		list[2] = "If the vehcile has a tow hook installed in the rear select Yes. If not, select No and specify the front tie down location for dyno installation."
	return list